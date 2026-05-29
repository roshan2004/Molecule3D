"""Tabular molecule libraries: read CSV/XLSX, compute descriptors, pick a diverse subset.

This is the one place MolScope works on a *table of molecules* (rows with columns
such as an id, a SMILES string, and numeric properties) rather than a single 3D
structure. It supports a common library-prep task: take a spreadsheet of
candidates and pick ``n`` that are spread out in descriptor space.

Selection runs on numeric descriptors. Those can be columns already in the table
(e.g. ``MW``, ``ALogP``) or computed from a SMILES column with RDKit (Crippen
``MolLogP`` is RDKit's standard stand-in for ALogP). The picker is a pure-NumPy
MaxMin farthest-first traversal, so selection itself needs no optional backend;
only XLSX reading (``openpyxl``) and SMILES descriptors (``rdkit``) do.
"""

from __future__ import annotations

import csv
import os
from dataclasses import dataclass
from typing import Optional

import numpy as np

#: RDKit descriptors computed from SMILES when none are named. ``MolLogP`` is
#: RDKit's Crippen logP, the usual ALogP equivalent.
DEFAULT_SMILES_DESCRIPTORS = (
    "MolWt",
    "MolLogP",
    "TPSA",
    "NumHDonors",
    "NumHAcceptors",
    "NumRotatableBonds",
)


@dataclass
class MoleculeTable:
    """A simple in-memory table of molecules: ordered columns and row dicts."""

    columns: list[str]
    rows: list[dict]

    def __len__(self) -> int:
        return len(self.rows)

    def column(self, name: str) -> list:
        """Return the raw values of one column, or raise if it is absent."""
        if name not in self.columns:
            raise KeyError(
                f"column {name!r} not found; available columns: {', '.join(self.columns)}"
            )
        return [row.get(name) for row in self.rows]

    def numeric_matrix(self, names: list[str]) -> np.ndarray:
        """Return an ``(n_rows, len(names))`` float matrix from the named columns.

        Non-numeric or empty cells become ``NaN`` so rows with missing values can
        be excluded by the selector rather than crashing the parse.
        """
        if not names:
            raise ValueError("no descriptor columns given")
        cols = [self.column(name) for name in names]
        matrix = np.full((len(self.rows), len(names)), np.nan, dtype=float)
        for j, values in enumerate(cols):
            for i, value in enumerate(values):
                matrix[i, j] = _to_float(value)
        return matrix

    def with_columns(self, names: list[str], matrix: np.ndarray) -> MoleculeTable:
        """Return a copy with extra numeric columns appended (e.g. computed descriptors)."""
        new_columns = list(self.columns) + [n for n in names if n not in self.columns]
        new_rows = []
        for i, row in enumerate(self.rows):
            merged = dict(row)
            for j, name in enumerate(names):
                merged[name] = float(matrix[i, j])
            new_rows.append(merged)
        return MoleculeTable(columns=new_columns, rows=new_rows)

    def select_rows(self, indices) -> MoleculeTable:
        """Return a new table with only the given row indices, in that order."""
        return MoleculeTable(columns=list(self.columns), rows=[self.rows[i] for i in indices])

    def write(self, path: str) -> None:
        """Write the table to ``.csv`` or ``.xlsx`` based on the extension."""
        ext = _extension(path)
        if ext in (".csv", ".tsv"):
            _write_csv(self, path, delimiter="\t" if ext == ".tsv" else ",")
        elif ext in (".xlsx", ".xlsm"):
            _write_xlsx(self, path)
        else:
            raise ValueError(f"unsupported output extension {ext!r}; use .csv or .xlsx")


def read_table(path: str, *, sheet: Optional[str] = None) -> MoleculeTable:
    """Read a molecule table from ``.csv``/``.tsv`` or ``.xlsx``/``.xlsm``/``.xls``.

    The first row is treated as the header. XLSX reading needs the ``openpyxl``
    package (``pip install "molscope[xlsx]"``); CSV reading uses the standard
    library only. ``sheet`` selects a worksheet by name for spreadsheet inputs
    (default: the first sheet).
    """
    ext = _extension(path)
    if ext in (".csv", ".tsv"):
        return _read_csv(path, delimiter="\t" if ext == ".tsv" else ",")
    if ext in (".xlsx", ".xlsm", ".xls"):
        return _read_xlsx(path, sheet=sheet)
    raise ValueError(
        f"unsupported table extension {ext!r}; use .csv, .tsv, or .xlsx"
    )


def smiles_descriptors(
    smiles: list, names: Optional[list[str]] = None
) -> tuple[np.ndarray, list[str]]:
    """Compute RDKit descriptors for a list of SMILES strings.

    Returns ``(matrix, names)`` where ``matrix`` is ``(len(smiles), len(names))``.
    Unparseable or empty SMILES yield a row of ``NaN`` rather than an error.
    Needs RDKit (``pip install "molscope[chem]"``).
    """
    from rdkit import RDLogger

    from .chem import _require_rdkit

    Chem, _ = _require_rdkit()
    from rdkit.Chem import Descriptors

    selected = list(names) if names else list(DEFAULT_SMILES_DESCRIPTORS)
    desc_map = dict(Descriptors._descList)
    unknown = [n for n in selected if n not in desc_map]
    if unknown:
        raise ValueError(f"unknown RDKit descriptor(s): {', '.join(unknown)}")

    # Silence RDKit's per-molecule parse errors: in a batch, unparseable SMILES
    # are expected and become NaN rows, not a wall of stderr noise.
    RDLogger.DisableLog("rdApp.*")
    try:
        matrix = np.full((len(smiles), len(selected)), np.nan, dtype=float)
        for i, smi in enumerate(smiles):
            if not smi or not str(smi).strip():
                continue
            mol = Chem.MolFromSmiles(str(smi))
            if mol is None:
                continue
            for j, name in enumerate(selected):
                try:
                    matrix[i, j] = float(desc_map[name](mol))
                except Exception:
                    matrix[i, j] = np.nan
    finally:
        RDLogger.EnableLog("rdApp.*")
    return matrix, selected


def select_diverse(matrix, n: int, *, standardize: bool = True) -> list[int]:
    """Pick ``n`` diverse rows by MaxMin (farthest-first) selection.

    ``matrix`` is an ``(n_rows, n_features)`` array of descriptors. Rows with any
    missing (``NaN``) feature are excluded from the candidate pool. When
    ``standardize`` is true (default), features are z-scored first so no single
    descriptor dominates the distance purely because of its scale.

    Returns the chosen row indices (into the original ``matrix``), seeded from the
    most extreme point and then greedily maximising the minimum distance to the
    already-chosen set. Selection is deterministic. If fewer valid rows than ``n``
    exist, all of them are returned.
    """
    if n <= 0:
        raise ValueError("n must be a positive integer")
    data = np.asarray(matrix, dtype=float)
    if data.ndim != 2:
        raise ValueError("matrix must be 2-dimensional (n_rows, n_features)")

    valid = ~np.isnan(data).any(axis=1)
    original = np.nonzero(valid)[0]
    points = data[valid]
    if len(points) == 0:
        raise ValueError("no rows with complete descriptors to select from")

    if standardize:
        mean = points.mean(axis=0)
        std = points.std(axis=0)
        std[std == 0] = 1.0
        points = (points - mean) / std

    k = min(n, len(points))
    # Seed with the point farthest from the centroid: deterministic and extreme.
    centroid = points.mean(axis=0)
    first = int(np.argmax(np.linalg.norm(points - centroid, axis=1)))

    chosen = [first]
    min_dist = np.linalg.norm(points - points[first], axis=1)
    while len(chosen) < k:
        min_dist[chosen] = -np.inf  # never re-pick a chosen row
        nxt = int(np.argmax(min_dist))
        chosen.append(nxt)
        min_dist = np.minimum(min_dist, np.linalg.norm(points - points[nxt], axis=1))
    return [int(original[i]) for i in chosen]


# -- internal helpers -------------------------------------------------------

def _extension(path: str) -> str:
    return os.path.splitext(os.fspath(path))[1].lower()


def _to_float(value) -> float:
    if value is None:
        return float("nan")
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return float("nan")
    try:
        return float(text)
    except ValueError:
        return float("nan")


def _read_csv(path: str, delimiter: str = ",") -> MoleculeTable:
    with open(path, newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle, delimiter=delimiter)
        columns = list(reader.fieldnames or [])
        if not columns:
            raise ValueError(f"{path}: no header row found")
        rows = [dict(row) for row in reader]
    return MoleculeTable(columns=columns, rows=rows)


def _read_xlsx(path: str, sheet: Optional[str] = None) -> MoleculeTable:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - exercised via the error-path test
        raise ImportError(
            "reading .xlsx files needs openpyxl; install with: pip install 'molscope[xlsx]'"
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet] if sheet else workbook[workbook.sheetnames[0]]
    rows_iter = worksheet.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        workbook.close()
        raise ValueError(f"{path}: worksheet is empty") from None
    columns = [str(c) if c is not None else f"col{i}" for i, c in enumerate(header)]
    rows = []
    for raw in rows_iter:
        if raw is None or all(cell is None for cell in raw):
            continue
        rows.append({columns[i]: raw[i] if i < len(raw) else None for i in range(len(columns))})
    workbook.close()
    return MoleculeTable(columns=columns, rows=rows)


def _write_csv(table: MoleculeTable, path: str, delimiter: str = ",") -> None:
    with open(path, "w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=table.columns, delimiter=delimiter)
        writer.writeheader()
        for row in table.rows:
            writer.writerow({col: row.get(col, "") for col in table.columns})


def _write_xlsx(table: MoleculeTable, path: str) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as exc:  # pragma: no cover - exercised via the error-path test
        raise ImportError(
            "writing .xlsx files needs openpyxl; install with: pip install 'molscope[xlsx]'"
        ) from exc

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(table.columns)
    for row in table.rows:
        worksheet.append([row.get(col) for col in table.columns])
    workbook.save(path)
